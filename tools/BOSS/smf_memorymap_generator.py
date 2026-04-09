"""
Memory Map Header Generator Module
Excelファイルからメモリマップヘッダーファイルを生成するためのモジュール
"""

import pandas as pd
from datetime import datetime


def generate_memory_map_header(excel_file):
    """Excelファイルからメモリマップヘッダーファイルを生成"""
    try:
        import time
        max_retries = 3
        retry_delay = 1
        
        for attempt in range(max_retries):
            try:
                # openpyxlで読み取り専用モードを試行
                from openpyxl import load_workbook
                print(f"Attempt {attempt + 1}: Opening Excel file in read-only mode...")
                wb = load_workbook(excel_file, read_only=True, data_only=True)
                available_sheets = wb.sheetnames
                print(f"Available sheets: {available_sheets}")
                
                # データを読み込む
                excel_data_success = True
                break
                
            except PermissionError as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    print(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                    continue
                else:
                    raise e
            except Exception as e:
                print(f"Read-only mode failed: {e}")
                # フォールバック: pandasで試行
                xl_file = pd.ExcelFile(excel_file, engine='openpyxl')
                available_sheets = xl_file.sheet_names
                excel_data_success = True
                break
        
        header_content = f"""#ifndef SMF_MEMORY_MAP_H
#define SMF_MEMORY_MAP_H

/*
 * Auto-generated Memory Map Header
 * Generated from: {excel_file.split('\\')[-1]}
 * Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
 */

// Base Memory Configuration
#define SMF_SIZE 0x08000000 // 128MB
#define SMF_START_ADDRESS 0x00000000
#define SMF_END_ADDRESS   0x07FFFFFF

"""
        
        # メモリマップシートを処理
        memory_map_sheet = None
        for sheet_name in available_sheets:
            if 'メモリマップ' in sheet_name:
                memory_map_sheet = sheet_name
                break
        
        if memory_map_sheet:
            print(f"\n=== Processing Sheet: {memory_map_sheet} ===")
            df = pd.read_excel(excel_file, sheet_name=memory_map_sheet, engine='openpyxl')
            print(f"Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            
            # ヘッダー行を見つける
            header_row = _find_header_row(df)
            
            if header_row is not None:
                # デバイスマッピングを抽出
                device_mappings = _extract_device_mappings(df, header_row)
                
                # Device IDセクションを生成
                header_content += _generate_device_id_section(memory_map_sheet, device_mappings)
                
                # Data IDセクションを生成
                header_content += _generate_data_id_section(df, header_row, memory_map_sheet, device_mappings)
                
                # Memory Address Mappingセクションを生成
                header_content += _generate_memory_address_section(df, header_row, memory_map_sheet, device_mappings)
                
            else:
                print("❌ Header row not found in memory map sheet")
        else:
            print("❌ Memory map sheet not found")
        
        header_content += "\n#endif // SMF_MEMORY_MAP_H\n"
        
        print("\n=== Final Generated Header Content ===")
        print(header_content)
        
        return header_content
        
    except Exception as e:
        print(f"❌ Excelファイル読み込みエラー: {e}")
        print("❌ Excelファイルにアクセスできません。ファイルが開いていないか確認してください。")
        print("   - Excelファイルを閉じてから再実行してください")
        print("   - ファイルの権限を確認してください")
        print("   - ネットワーク接続を確認してください")
        return None


def _find_header_row(df):
    """ヘッダー行を見つける"""
    for i in range(min(10, len(df))):
        row_str = ' '.join([str(df.iloc[i][col]) for col in df.columns if pd.notna(df.iloc[i][col])])
        if 'System' in row_str and ('Device' in row_str or 'ID' in row_str):
            print(f"Found header row at index {i}")
            return i
    return None


def _extract_device_mappings(df, header_row):
    """デバイスマッピングを抽出"""
    print(f"\nHeader row content:")
    for j, col in enumerate(df.columns):
        if pd.notna(df.iloc[header_row, j]):
            print(f"  Column {j}: {df.iloc[header_row, j]}")
    
    print(f"\nProcessing data rows {header_row + 1} to {len(df)}:")
    
    device_mappings = {}  # {upper_nibble: device_name} の辞書
    
    # DataIDとSystem列からデバイス名とIDのマッピングを抽出
    for i in range(header_row + 1, len(df)):
        try:
            # DataID列（D列、インデックス3）からデバイスIDを取得
            if len(df.columns) > 3 and pd.notna(df.iloc[i, 3]):
                data_id_raw = str(df.iloc[i, 3]).strip()
                if data_id_raw and data_id_raw != 'nan' and data_id_raw != '-':
                    # DataIDの上位ニブル（0x1, 0x2, 0xC など）を抽出
                    if len(data_id_raw) >= 3:  # "0xX" の最小長
                        upper_nibble = data_id_raw[:3]  # "0x1", "0x2", "0xC" など
                        
                        device_name = _extract_device_name(df, i)
                        
                        # デバイス名が取得できた場合のみマッピングに追加
                        if device_name and upper_nibble not in device_mappings:
                            device_mappings[upper_nibble] = device_name
                            print(f"      Debug: Added mapping {upper_nibble} -> {device_name}")
        except Exception:
            continue
    
    return device_mappings


def _extract_device_name(df, row_index):
    """行からデバイス名を抽出"""
    device_name = None
    
    # まずAssignment (Device ID)列（C列、インデックス2）を確認
    if len(df.columns) > 2 and pd.notna(df.iloc[row_index, 2]):
        assignment_name = str(df.iloc[row_index, 2]).strip().upper()
        if assignment_name and assignment_name != 'NAN':
            # "MAIN PIC\n(1)" や "COM PIC\n(2)" などから "MAIN", "COM" を抽出
            device_name = assignment_name.replace(" PIC", "").replace(" MCU", "")
            device_name = device_name.split('\n')[0].strip()  # 改行で分割して最初の部分のみ
            device_name = ''.join(c for c in device_name if c.isalpha())  # アルファベットのみ
            print(f"      Debug: Row {row_index}, Assignment={assignment_name} -> {device_name}")
    
    # Assignment列が空の場合、System列（B列、インデックス1）を確認
    if not device_name and len(df.columns) > 1 and pd.notna(df.iloc[row_index, 1]):
        system_name = str(df.iloc[row_index, 1]).strip().upper()
        if system_name and system_name != 'NAN':
            device_name = system_name.replace(" PIC", "").replace(" MCU", "")
            device_name = device_name.split('\n')[0].strip()  # 改行で分割して最初の部分のみ
            device_name = ''.join(c for c in device_name if c.isalpha())  # アルファベットのみ
            print(f"      Debug: Row {row_index}, System={system_name} -> {device_name}")
    
    return device_name


def _generate_device_id_section(memory_map_sheet, device_mappings):
    """Device IDセクションを生成"""
    content = f"\n// ===== Device IDs from {memory_map_sheet} =====\n"
    
    for upper_nibble, device_name in sorted(device_mappings.items()):
        device_id = upper_nibble + "0"  # "0x1" -> "0x10"
        content += f"#define {device_name}_DEVICE_ID {device_id}\n"
        print(f"    Added: {device_name}_DEVICE_ID -> {device_id}")
    
    return content


def _generate_data_id_section(df, header_row, memory_map_sheet, device_mappings):
    """Data IDセクションを生成"""
    content = f"\n// ===== Data IDs from {memory_map_sheet} =====\n"
    
    # DataIDを抽出
    data_ids = set()  # 重複を避けるためにsetを使用
    
    for i in range(header_row + 1, len(df)):
        try:
            # DataID列（D列、インデックス3）からDataIDを取得
            if len(df.columns) > 3 and pd.notna(df.iloc[i, 3]):
                data_id_raw = str(df.iloc[i, 3]).strip()
                if data_id_raw and data_id_raw != 'nan' and data_id_raw != '-':
                    # Sub Assignment列（E列、インデックス4）から説明を取得
                    sub_assignment = ""
                    if len(df.columns) > 4 and pd.notna(df.iloc[i, 4]):
                        sub_assignment = str(df.iloc[i, 4]).strip()
                    
                    # DataIDとSub Assignmentのペアを保存
                    data_ids.add((data_id_raw, sub_assignment))
        except Exception:
            continue
    
    # 抽出されたDataIDを出力
    for data_id, sub_assignment in sorted(data_ids):
        definition_name = _create_data_id_definition_name(data_id, sub_assignment, device_mappings)
        content += f"#define {definition_name} {data_id}\n"
        print(f"    Added: {definition_name} -> {data_id} ({sub_assignment})")
    
    return content


def _create_data_id_definition_name(data_id, sub_assignment, device_mappings):
    """DataID用の定義名を作成"""
    # デバイスプレフィックスを決定
    device_prefix = "UNKNOWN"
    for upper_nibble, device_name in device_mappings.items():
        if data_id.startswith(upper_nibble):
            device_prefix = device_name
            break
    
    # Sub Assignmentをクリーンアップしてdefine名を作成
    if sub_assignment:
        sub_clean = sub_assignment.upper().strip()
        # デバイス名の重複を除去
        device_id_patterns = [
            " (1)", " (2)", " (6)", " (7)", " (8)", " (9)", " (B)", " (C)",
            "(1) ", "(2) ", "(6) ", "(7) ", "(8) ", "(9) ", "(B) ", "(C) ",
            "MAIN ", "COM ", "APRS ", "CAM ", "CHO ", "SATO ", "BHU ", "CIGS "
        ]
        for pattern in device_id_patterns:
            sub_clean = sub_clean.replace(pattern, " ").strip()
        
        # 安全な定義名に変換
        safe_name = sub_clean.replace(' ', '_').replace('-', '_').replace('&', '_')
        safe_name = safe_name.replace('(', '').replace(')', '').replace(',', '_')
        safe_name = ''.join(c for c in safe_name if c.isalnum() or c == '_')
        while '__' in safe_name:
            safe_name = safe_name.replace('__', '_')
        safe_name = safe_name.strip('_')
        
        if safe_name:
            return f"ID_{device_prefix}_{safe_name}"
        else:
            return f"ID_{device_prefix}_{data_id.replace('0x', '')}"
    else:
        return f"ID_{device_prefix}_{data_id.replace('0x', '')}"


def _generate_memory_address_section(df, header_row, memory_map_sheet, device_mappings):
    """Memory Address Mappingセクションを生成"""
    content = f"\n// ===== Memory Address Mappings from {memory_map_sheet} =====\n"
    
    # 重複チェック用のセット
    generated_definitions = set()
    
    for i in range(header_row + 1, len(df)):
        try:
            sub_assignment = None
            start_address = None
            end_address = None
            data_id = None
            
            # Sub Assignment列（E列、インデックス4）
            if len(df.columns) > 4 and pd.notna(df.iloc[i, 4]):
                sub_assignment = str(df.iloc[i, 4]).strip()
            
            # Start address列（I列、インデックス8）
            if len(df.columns) > 8 and pd.notna(df.iloc[i, 8]):
                start_address = str(df.iloc[i, 8]).strip()
            
            # End address列（J列、インデックス9）
            if len(df.columns) > 9 and pd.notna(df.iloc[i, 9]):
                end_address = str(df.iloc[i, 9]).strip()
            
            # DataID列（D列、インデックス3）からデバイスIDを取得
            if len(df.columns) > 3 and pd.notna(df.iloc[i, 3]):
                data_id_raw = str(df.iloc[i, 3]).strip()
                if data_id_raw and data_id_raw != 'nan' and data_id_raw != '-':
                    data_id = data_id_raw
            
            # デバッグ情報を表示（CIGS関連のみ）
            if i < header_row + 90 and (data_id and "0xC" in data_id):
                print(f"    Row {i}: DataID={data_id}, Sub={sub_assignment}, Start={start_address}, End={end_address}")
                print(f"      *** CIGS RELATED ***")
            
            # データが有効な場合のみ処理
            if sub_assignment and start_address and end_address and data_id:
                definition_name = _create_address_definition_name(data_id, sub_assignment, device_mappings)
                
                if definition_name:
                    # アドレスを16進数形式に正規化
                    if not start_address.startswith('0x') and start_address.isalnum():
                        start_address = f"0x{start_address.upper()}"
                    if not end_address.startswith('0x') and end_address.isalnum():
                        end_address = f"0x{end_address.upper()}"
                    
                    # 重複チェック - 同じ定義名が既に存在する場合はスキップ
                    start_def = f"{definition_name}_START_ADDRESS"
                    end_def = f"{definition_name}_END_ADDRESS"
                    
                    if start_def not in generated_definitions:
                        # 定義を追加
                        content += f"#define {start_def} {start_address}\n"
                        content += f"#define {end_def} {end_address}\n"
                        generated_definitions.add(start_def)
                        generated_definitions.add(end_def)
                        print(f"    Added: {definition_name} -> {start_address} to {end_address}")
                    else:
                        print(f"    Skipped duplicate: {definition_name}")
        except Exception as e:
            continue
    
    return content


def _create_address_definition_name(data_id, sub_assignment, device_mappings):
    """アドレス用の定義名を作成"""
    sub_clean = sub_assignment.upper().strip()
    
    # Sub AssignmentからDevice IDを除去（デバイス名の重複のみ）
    device_id_patterns = [
        " (1)", " (2)", " (6)", " (7)", " (8)", " (9)", " (B)", " (C)",
        "(1) ", "(2) ", "(6) ", "(7) ", "(8) ", "(9) ", "(B) ", "(C) ",
        "MAIN ", "COM ", "APRS ", "CAM ", "CHO ", "SATO ", "BHU ", "CIGS "
    ]
    for pattern in device_id_patterns:
        sub_clean = sub_clean.replace(pattern, " ").strip()
    
    # デバイスプレフィックスをDataIDから決定（動的）
    device_prefix = None
    for upper_nibble, device_name in device_mappings.items():
        if data_id.startswith(upper_nibble):  # "0x1", "0x2", "0xC" などの比較
            device_prefix = device_name
            break
    
    if not device_prefix:
        return None
    
    # Sub Assignment文字列をそのままdefine名に変換
    # スペース、ハイフン、アンパサンドをアンダースコアに変換
    safe_name = sub_clean.replace(' ', '_').replace('-', '_').replace('&', '_')
    safe_name = safe_name.replace('(', '').replace(')', '').replace(',', '_')
    
    # 不要な文字を除去し、連続するアンダースコアを1つにまとめる
    safe_name = ''.join(c for c in safe_name if c.isalnum() or c == '_')
    while '__' in safe_name:
        safe_name = safe_name.replace('__', '_')
    safe_name = safe_name.strip('_')
    
    # 最終的な定義名を作成
    if safe_name:
        return f"{device_prefix}_{safe_name}"
    else:
        return None
